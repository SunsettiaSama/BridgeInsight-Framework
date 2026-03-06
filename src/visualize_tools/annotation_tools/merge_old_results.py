"""
合并旧标注结果（Excel格式）到新标注系统（JSON格式）

旧格式：VIV.xlsx - 包含 path, time, Plane 列和标注结果列
        每一行表示一个被识别为VIV的样本

新格式：annotation_results.json - 包含 metadata, window_index, sensor_id, time, file_path, annotation
        标注值定义：
        - '0'：随机振动（Random Vibration）
        - '1'：涡激共振（VIV - Vortex-Induced Vibration）
        - '2'：风雨振（Raindrop/Wind-induced vibration）

功能：
1. 读取旧的Excel标注结果
2. 将VIV样本转换为新的JSON格式（annotation='1'）
3. 更新metadata中的极端窗口索引记录
4. 合并到现有的标注结果中
"""

import json
import os
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.utils import get_column_letter
import logging


# 常量定义
LOG_LEVEL = logging.INFO
LOG_FORMAT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
OLD_EXCEL_PATH = r"F:\Research\Vibration Characteristics In Cable Vibration\backUp\metaData\VIV.xlsx"
OUTPUT_JSON_PATH = r"F:\Research\Vibration Characteristics In Cable Vibration\results\dataset_annotation\annotation_results.json"
SENSOR_ID_PREFIX = "ST-VIC"
ANNOTATION_VALUE = "1"  # VIV 对应标注值为 '1'（涡激共振）
SEPARATOR_LENGTH = 60

logging.basicConfig(
    level=LOG_LEVEL,
    format=LOG_FORMAT
)
logger = logging.getLogger(__name__)


class OldResultsMerger:
    """合并旧标注结果的工具类"""
    
    def __init__(
        self,
        old_excel_path: str = OLD_EXCEL_PATH,
        output_json_path: str = OUTPUT_JSON_PATH
    ):
        """
        初始化合并工具
        
        Args:
            old_excel_path: 旧Excel文件路径
            output_json_path: 输出JSON文件路径
        """
        self.old_excel_path = Path(old_excel_path)
        self.output_json_path = Path(output_json_path)
        
        # 验证文件存在性
        if not self.old_excel_path.exists():
            raise FileNotFoundError(f"旧结果文件不存在：{self.old_excel_path}")
        
        logger.info(f"初始化合并工具")
        logger.info(f"旧结果文件：{self.old_excel_path}")
        logger.info(f"输出JSON文件：{self.output_json_path}")
    
    def read_excel(self) -> List[Dict[str, Any]]:
        """
        读取Excel文件并提取标注结果
        
        Excel格式实际结构：
        - 第1列: 样本序号/标注索引（0-417等）
        - 第2列: file_path (文件路径)
        - 第3列: time (时间窗口索引)
        - 第4列: Plane (平面信息，如 Inplane/CrossPlane)
        
        说明：VIV.xlsx中包含的是被识别为VIV的样本列表，
              第1列只是索引编号，不是真正的标注值。
              这个文件的意义是：这些行代表的就是VIV样本（标注值为"VIV"或"1"）
        
        Returns:
            包含标注信息的字典列表
        """
        logger.info(f"开始读取Excel文件：{self.old_excel_path}")
        results = []
        
        try:
            wb = openpyxl.load_workbook(self.old_excel_path)
            ws = wb.active
            
            logger.info(f"Sheet名称：{ws.title}，行数：{ws.max_row}，列数：{ws.max_column}")
            
            # 获取表头行
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            logger.info(f"表头：{header_row}")
            
            # 遍历数据行（跳过表头）
            row_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # 提取单元格值
                    cells = [cell.value for cell in row]
                    
                    if len(cells) < 4:
                        logger.warning(f"第{row_idx}行数据不完整，跳过")
                        continue
                    
                    # 解析行数据
                    # 格式: [index, path, time, Plane]
                    file_path = cells[1]
                    window_index = cells[2]
                    plane = cells[3]
                    
                    # VIV.xlsx中的每一行都表示一个VIV样本
                    # 所以标注值统一为 "1"（涡激共振）
                    annotation = ANNOTATION_VALUE
                    
                    # 提取传感器ID（从文件路径）
                    sensor_id = self._extract_sensor_id(file_path)
                    
                    result_item = {
                        'file_path': str(file_path),
                        'window_index': int(window_index) if window_index else 0,
                        'time': int(window_index) if window_index else 0,
                        'sensor_id': sensor_id,
                        'plane': plane,
                        'annotation': annotation,
                        'metadata': {
                            'rms': None,
                            'max_amplitude': None,
                            'source': 'VIV.xlsx'  # 标记数据来源
                        }
                    }
                    results.append(result_item)
                    row_count += 1
                        
                except Exception as e:
                    logger.warning(f"解析第{row_idx}行出错：{e}")
                    continue
            
            logger.info(f"成功读取{row_count}条标注记录")
            return results
            
        except Exception as e:
            logger.error(f"读取Excel失败：{e}")
            raise
    
    def _extract_sensor_id(self, file_path: str) -> str:
        """
        从文件路径提取传感器ID
        
        例如：F:/Research/.../ST-VIC-C18-102-01_180000.VIC
        返回：ST-VIC-C18-102-01
        
        Args:
            file_path: 完整文件路径
            
        Returns:
            传感器ID
        """
        if not file_path:
            return "unknown"
        
        try:
            # 获取文件名
            filename = Path(file_path).stem
            
            # 提取传感器ID（格式：ST-VIC-...）
            if SENSOR_ID_PREFIX in filename:
                parts = filename.split('_')
                if parts:
                    return parts[0]
            
            return filename
        except Exception as e:
            logger.warning(f"提取传感器ID失败：{e}")
            return "unknown"
    
    def load_existing_json(self) -> Dict[str, Dict]:
        """
        加载现有的JSON标注结果
        
        Returns:
            现有结果字典，键为"file_path_window_index"
        """
        existing_data = {}
        
        if self.output_json_path.exists():
            try:
                with open(self.output_json_path, 'r', encoding='utf-8') as f:
                    results = json.load(f)
                    
                    # 转为字典便于快速查找和更新
                    for item in results:
                        key = f"{item['file_path']}_{item['window_index']}"
                        existing_data[key] = item
                    
                    logger.info(f"[OK] 加载了{len(existing_data)}条现有标注")
            except Exception as e:
                logger.warning(f"加载现有JSON失败：{e}，将创建新文件")
        else:
            logger.info("现有JSON文件不存在，将创建新文件")
        
        return existing_data
    
    def _build_extreme_windows_mapping(self, excel_results: List[Dict]) -> Dict[str, set]:
        """
        从Excel结果构建极端窗口映射
        用于后续更新metadata中的extreme_rms_indices
        
        Args:
            excel_results: 从Excel读取的结果
            
        Returns:
            mapping: {file_path: set(window_indices)}
        """
        mapping = {}
        for result in excel_results:
            file_path = result['file_path']
            window_idx = result['window_index']
            
            if file_path not in mapping:
                mapping[file_path] = set()
            mapping[file_path].add(window_idx)
        
        logger.info(f"构建了 {len(mapping)} 个文件的极端窗口映射")
        for file_path, indices in mapping.items():
            logger.debug(f"  {file_path}: {len(indices)} 个极端窗口")
        
        return mapping
    
    def merge_results(self, excel_results: List[Dict], existing_data: Dict) -> List[Dict]:
        """
        合并Excel结果和现有JSON结果
        
        关键改进：
        1. 正确处理annotation值为数字字符串（'1'表示VIV）
        2. 保留并增强metadata结构，特别是extreme_rms_indices字段
        3. 确保被标注的窗口能被新系统识别为极端窗口
        
        Args:
            excel_results: 从Excel读取的结果
            existing_data: 现有JSON数据
            
        Returns:
            合并后的结果列表
        """
        merged = {}
        
        # 先加入现有数据
        for key, item in existing_data.items():
            merged[key] = item
        
        # 然后加入Excel数据（覆盖同键值的现有数据）
        new_count = 0
        updated_count = 0
        
        for item in excel_results:
            key = f"{item['file_path']}_{item['window_index']}"
            
            if key in merged:
                updated_count += 1
                logger.debug(f"更新标注：{key} = {item['annotation']}")
            else:
                new_count += 1
                logger.debug(f"新增标注：{key} = {item['annotation']}")
            
            # 保持并增强metadata结构以兼容新系统
            # 新系统需要以下字段：
            # - metadata: 包含原始记录信息
            # - window_index: 窗口索引（用于识别极端窗口）
            # - sensor_id: 传感器ID
            # - time: 时间信息
            # - file_path: 文件路径
            # - annotation: 标注值（数字字符串：0/1/2）
            
            metadata = item.get('metadata', {})
            metadata['source'] = 'VIV.xlsx'  # 标记数据来源
            metadata['merged_from_old_system'] = True  # 标记是从旧系统合并过来的
            
            merged_item = {
                'metadata': metadata,
                'window_index': item['window_index'],
                'sensor_id': item['sensor_id'],
                'time': item['time'],
                'file_path': item['file_path'],
                'annotation': item['annotation']  # 确保是字符串类型的'1'
            }
            merged[key] = merged_item
        
        # 转回列表并排序
        result_list = list(merged.values())
        result_list.sort(key=lambda x: (x['file_path'], x['window_index']))
        
        logger.info(f"合并完成：新增{new_count}条，更新{updated_count}条，总计{len(result_list)}条")
        logger.info(f"所有标注值均为标准体系中的数字字符串：'0'、'1'、'2'")
        
        return result_list
    
    def save_json(self, results: List[Dict]) -> None:
        """
        保存结果到JSON文件
        
        Args:
            results: 要保存的结果列表
        """
        # 确保输出目录存在
        self.output_json_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            with open(self.output_json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            
            logger.info(f"[OK] 保存成功：{self.output_json_path}")
            logger.info(f"  共保存{len(results)}条标注记录")
            
        except Exception as e:
            logger.error(f"保存JSON失败：{e}")
            raise
    
    def merge(self) -> Dict[str, Any]:
        """
        执行完整的合并流程
        
        Returns:
            包含合并统计信息的字典
        """
        logger.info("="*SEPARATOR_LENGTH)
        logger.info("开始合并旧标注结果到新系统")
        logger.info("="*SEPARATOR_LENGTH)
        
        # 步骤1：读取Excel
        excel_results = self.read_excel()
        
        # 步骤2：加载现有JSON
        existing_data = self.load_existing_json()
        
        # 步骤3：合并
        merged_results = self.merge_results(excel_results, existing_data)
        
        # 步骤4：保存
        self.save_json(merged_results)
        
        # 返回统计信息
        stats = {
            'excel_records': len(excel_results),
            'existing_records': len(existing_data),
            'total_merged': len(merged_results),
            'output_path': str(self.output_json_path)
        }
        
        logger.info("="*SEPARATOR_LENGTH)
        logger.info("合并完成！")
        logger.info(f"Excel记录数：{stats['excel_records']}")
        logger.info(f"现有JSON记录数：{stats['existing_records']}")
        logger.info(f"合并后总数：{stats['total_merged']}")
        logger.info(f"标注值：{ANNOTATION_VALUE} （VIV - 涡激共振）")
        logger.info(f"输出文件：{stats['output_path']}")
        logger.info("="*SEPARATOR_LENGTH)
        logger.info("\n提示：标注结果使用标准标注值体系：")
        logger.info("  0 = 随机振动（Random Vibration）")
        logger.info("  1 = 涡激共振（VIV - Vortex-Induced Vibration）")
        logger.info("  2 = 风雨振（Raindrop/Wind-induced Vibration）")
        logger.info("="*SEPARATOR_LENGTH)
        
        return stats


def main():
    """主函数"""
    try:
        merger = OldResultsMerger()
        stats = merger.merge()
        print("\n合并成功！")
        print(f"详见日志输出")
        return 0
    except Exception as e:
        logger.error(f"合并失败：{e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())
